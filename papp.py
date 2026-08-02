import streamlit as st
import pandas as pd
import os
from datetime import datetime
from streamlit_drawable_canvas import st_canvas
import io
from PIL import Image

# Importaciones para Excel
import openpyxl
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter

# Importaciones para PDF
from reportlab.lib.pagesizes import letter, landscape
from reportlab.pdfgen import canvas as pdf_canvas
from reportlab.platypus import Table as RLTable, TableStyle, Image as RLImage, Paragraph
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors

# ==========================================
# CONFIGURACIÓN Y ESTILOS (Colores LIF Brands)
# ==========================================
st.set_page_config(page_title="Control de Recepción - LIF Brands", layout="wide", initial_sidebar_state="collapsed")

st.markdown("""
<style>
/* Forzar colores claros y tipografía sin romper iconos internos (span) */
.stApp { 
    background-color: #F8FAF9 !important; 
}

html, body, p, label, div { 
    font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif; 
    color: #1f2937; 
}

h1, h2, h3, h4, h5, h6 { 
    color: #115e59 !important; 
    font-weight: 700 !important; 
}

/* Colores para las etiquetas de los inputs */
.stTextInput label, .stSelectbox label, .stDateInput label, .stNumberInput label, .stFileUploader label, .stTextArea label {
    color: #0f766e !important;
    font-weight: bold !important;
}

/* Arreglo para Expander (Ver detalles) */
[data-testid="stExpander"] details summary {
    background-color: #f0fdf4 !important;
    color: #115e59 !important;
    font-weight: bold;
    border-radius: 6px;
}
[data-testid="stExpander"] details summary p {
    color: #115e59 !important;
    font-weight: bold;
}

/* Botones Principales (Verde LIF) */
[data-testid="baseButton-primary"] {
    background-color: #115e59 !important;
    color: #FFFFFF !important;
    border: none !important;
    border-radius: 6px !important;
    font-weight: bold !important;
}

[data-testid="baseButton-primary"]:hover {
    background-color: #0f766e !important;
}

/* Botones Secundarios */
[data-testid="baseButton-secondary"] {
    background-color: #FFFFFF !important;
    color: #115e59 !important;
    border: 1px solid #115e59 !important;
    border-radius: 6px !important;
    font-weight: bold;
}

[data-testid="baseButton-secondary"]:hover {
    background-color: #f0fdf4 !important;
}

/* Tarjetas (Cards) de los registros */
.record-card {
    background-color: #ffffff;
    border: 1px solid #e2e8f0;
    border-radius: 8px;
    padding: 16px;
    margin-bottom: 5px;
    box-shadow: 0 1px 3px rgba(0,0,0,0.1);
    border-left: 5px solid #0f766e;
}

.record-header {
    font-size: 1.1em;
    font-weight: bold;
    color: #115e59;
    margin-bottom: 8px;
}

.record-sub {
    color: #475569;
    font-size: 0.9em;
}

.status-pendiente { color: #d97706; font-weight: bold; }
.status-aprobado { color: #16a34a; font-weight: bold; }
.status-rechazado { color: #dc2626; font-weight: bold; }

/* Borde visible para el área de dibujo de la firma */
canvas {
    border: 1px solid #94a3b8 !important;
    border-radius: 6px;
    cursor: crosshair;
}
</style>
""", unsafe_allow_html=True)

# ==========================================
# DIRECTORIOS Y FUNCIONES BASE
# ==========================================
FIRMAS_DIR = "firmas_recepcion"
EVIDENCIAS_DIR = "evidencias_recepcion"

for directorio in [FIRMAS_DIR, EVIDENCIAS_DIR]:
    if not os.path.exists(directorio):
        os.makedirs(directorio)

EXCEL_FILE = "registros_recepcion_coco.xlsx"

def mostrar_logo(ancho=160):
    if os.path.exists("logo.png"):
        st.image("logo.png", width=ancho)

def cargar_datos():
    if os.path.exists(EXCEL_FILE):
        return pd.read_excel(EXCEL_FILE, dtype={"ID_Registro": str})
    return pd.DataFrame()

def guardar_datos(df):
    df.to_excel(EXCEL_FILE, index=False)

def eliminar_registro(id_registro):
    df = cargar_datos()
    df = df[df["ID_Registro"] != str(id_registro)]
    guardar_datos(df)
    st.success(f"Registro #{id_registro} eliminado.")

def generar_id_registro():
    fecha_base = datetime.today().strftime("%Y%m%d")
    df = cargar_datos()
    
    if df.empty or "ID_Registro" not in df.columns:
        return fecha_base
    
    ids_existentes = df["ID_Registro"].astype(str).tolist()
    if fecha_base not in ids_existentes:
        return fecha_base
    
    contador = 1
    while f"{fecha_base}({contador})" in ids_existentes:
        contador += 1
    return f"{fecha_base}({contador})"

def generar_excel_bytes(df):
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Registros')
        worksheet = writer.sheets['Registros']
        
        max_row = worksheet.max_row
        max_col = worksheet.max_column
        
        if max_row > 1:
            ref = f"A1:{get_column_letter(max_col)}{max_row}"
            tab = Table(displayName="TablaRegistros", ref=ref)
            style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=True)
            tab.tableStyleInfo = style
            worksheet.add_table(tab)
            
            for col in worksheet.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                worksheet.column_dimensions[column].width = min(max_length + 2, 35)
                
    return output.getvalue()

# ==========================================
# GENERADOR DE PDF
# ==========================================
def generar_pdf_nuevo(registro):
    buffer = io.BytesIO()
    c = pdf_canvas.Canvas(buffer, pagesize=landscape(letter))
    width, height = landscape(letter)
    
    margin_x, margin_y = 40, 40
    usable_width = width - (2 * margin_x)
    col_widths = [140, 240, 60, 100, 50, usable_width - (140+240+60+100+50)]
    styles = getSampleStyleSheet()
    
    style_normal = ParagraphStyle('CellNormal', parent=styles['Normal'], fontName='Helvetica', fontSize=9, leading=11, textColor=colors.black)
    style_bold = ParagraphStyle('CellBold', parent=style_normal, fontName='Helvetica-Bold')
    style_center_bold = ParagraphStyle('CellCenterBold', parent=style_bold, alignment=1)
    style_meta = ParagraphStyle('CellMeta', parent=style_normal, fontSize=8, leading=10)

    logo_img = RLImage("logo.png", width=80, height=30) if os.path.exists("logo.png") else ""
        
    data = [
        [logo_img, Paragraph("Control de Recepción de Coco", style_center_bold), "", "", "", Paragraph(f"Código: R ICC/15-2<br/>Versión: 02<br/>ID: #{registro.get('ID_Registro', '')}", style_meta)],
        [Paragraph("Nombre del responsable:", style_bold), Paragraph(str(registro.get('Responsable', '')), style_normal), Paragraph("Fecha:", style_bold), Paragraph(str(registro.get('Fecha', '')), style_normal), Paragraph("Hora:", style_bold), Paragraph(str(registro.get('Hora', '')), style_normal)],
        [Paragraph("Descripción de materia prima:", style_bold), Paragraph(str(registro.get('Desc_Materia', '')), style_normal), Paragraph("Observaciones", style_center_bold), "", "", ""],
        [Paragraph("Nombre del proveedor", style_bold), Paragraph(str(registro.get('Proveedor', '')), style_normal), Paragraph(str(registro.get('Observaciones', 'Ninguna')), style_normal), "", "", ""],
        [Paragraph("Total de Fruta Ingresada Planta", style_bold), Paragraph(str(registro.get('Total_Fruta', '')), style_normal), "", "", "", ""],
        [Paragraph("Cantidad Unidades (Muestra)", style_bold), Paragraph(str(registro.get('Cant_Unidades', '')), style_normal), "", "", "", ""],
        [Paragraph("PARÁMETROS FISICOQUÍMICOS", style_center_bold), "", "", "", "", ""],
        [Paragraph("Muestra 1", style_center_bold), "", "", "", "", ""],
        [Paragraph("Unidades/Galón", style_normal), Paragraph(str(registro.get('unidades_galon_1', '')), style_normal), "", "", "", ""],
        [Paragraph("Volumen (Muestra)", style_normal), Paragraph(str(registro.get('volumen_1', '')), style_normal), "", "", "", ""],
        [Paragraph("Brix° (5 - 5.9)", style_normal), Paragraph(str(registro.get('brix_1', '')), style_normal), "", "", "", ""],
        [Paragraph("pH", style_normal), Paragraph(str(registro.get('ph_1', '')), style_normal), "", "", "", ""],
        [Paragraph("Muestra 2", style_center_bold), "", "", "", "", ""],
        [Paragraph("Unidades/Galón", style_normal), Paragraph(str(registro.get('unidades_galon_2', '')), style_normal), "", "", "", ""],
        [Paragraph("Volumen (Muestra)", style_normal), Paragraph(str(registro.get('volumen_2', '')), style_normal), "", "", "", ""],
        [Paragraph("Brix° (5 - 5.9)", style_normal), Paragraph(str(registro.get('brix_2', '')), style_normal), "", "", "", ""],
        [Paragraph("pH", style_normal), Paragraph(str(registro.get('ph_2', '')), style_normal), "", "", "", ""],
        [Paragraph("Muestra 3", style_center_bold), "", "", "", "", ""],
        [Paragraph("Unidades/Galón", style_normal), Paragraph(str(registro.get('unidades_galon_3', '')), style_normal), "", "", "", ""],
        [Paragraph("Volumen (Muestra)", style_normal), Paragraph(str(registro.get('volumen_3', '')), style_normal), "", "", "", ""],
        [Paragraph("Brix° (5 - 5.9)", style_normal), Paragraph(str(registro.get('brix_3', '')), style_normal), "", "", "", ""],
        [Paragraph("pH", style_normal), Paragraph(str(registro.get('ph_3', '')), style_normal), "", "", "", ""],
    ]
    
    style = TableStyle([
        ('GRID', (0,0), (-1,-1), 1, colors.black), ('BOX', (0,0), (-1,-1), 2, colors.black), ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('SPAN', (1,0), (4,0)), ('VALIGN', (1,0), (1,0), 'MIDDLE'), ('VALIGN', (5,0), (5,0), 'TOP'),
        ('SPAN', (2,2), (5,2)), ('SPAN', (2,3), (5,5)), ('VALIGN', (2,3), (5,5), 'TOP'),
        ('SPAN', (1,3), (1,3)), ('SPAN', (1,4), (1,4)), ('SPAN', (1,5), (1,5)),
        ('SPAN', (0,6), (-1,6)), ('SPAN', (0,7), (-1,7)), 
        ('SPAN', (1,8), (-1,8)), ('SPAN', (1,9), (-1,9)), ('SPAN', (1,10), (-1,10)), ('SPAN', (1,11), (-1,11)),
        ('SPAN', (0,12), (-1,12)), 
        ('SPAN', (1,13), (-1,13)), ('SPAN', (1,14), (-1,14)), ('SPAN', (1,15), (-1,15)), ('SPAN', (1,16), (-1,16)),
        ('SPAN', (0,17), (-1,17)), 
        ('SPAN', (1,18), (-1,18)), ('SPAN', (1,19), (-1,19)), ('SPAN', (1,20), (-1,20)), ('SPAN', (1,21), (-1,21)),
    ])
    
    row_heights = [45] + [22]*5 + [16]*16
    t = RLTable(data, colWidths=col_widths, rowHeights=row_heights)
    t.setStyle(style)
    
    w, h = t.wrapOn(c, usable_width, height)
    y_pos_table = height - margin_y - h
    t.drawOn(c, margin_x, y_pos_table)
    
    y_firma = y_pos_table - 45
    nombre_firma = registro.get("Firma_Jefe", "Sin firma")
    ruta_firma = os.path.join(FIRMAS_DIR, str(nombre_firma))
    
    if nombre_firma != "Sin firma" and os.path.exists(ruta_firma):
        c.drawImage(ruta_firma, width/2.0 - 60, y_firma, width=120, height=40, preserveAspectRatio=True, mask='auto')
        
    c.line(width/2.0 - 120, y_firma - 5, width/2.0 + 120, y_firma - 5)
    c.setFont("Helvetica-Bold", 9)
    c.drawCentredString(width/2.0, y_firma - 18, "Jefe de Calidad")
    
    c.save()
    buffer.seek(0)
    return buffer.getvalue()

# ==========================================
# CONTROL DE ESTADOS DE SESIÓN
# ==========================================
for state in ["nav_state", "form_logueado", "admin_logueado", "enviado_exitoso"]:
    if state not in st.session_state:
        st.session_state[state] = "home" if state == "nav_state" else False

# ==========================================
# 1. PANTALLA DE INICIO
# ==========================================
if st.session_state["nav_state"] == "home":
    col1, col2, col3 = st.columns([1, 2, 1])
    with col2:
        st.markdown("<br>", unsafe_allow_html=True)
        mostrar_logo(200)
        st.markdown("<h1 style='text-align: center; color: #115e59 !important;'>Control de Recepción de Coco</h1>", unsafe_allow_html=True)
        st.markdown("<p style='text-align: center; font-weight: bold;'>LIF Brands - Aseguramiento de Calidad</p>", unsafe_allow_html=True)
        st.markdown("<br>", unsafe_allow_html=True)
        
        if st.button("📝 Colaborador: Reportar nuevo ingreso", use_container_width=True, type="primary"):
            st.session_state["nav_state"] = "form_login"
            st.rerun()
            
        st.markdown("<br>", unsafe_allow_html=True)
        if st.button("🔒 Jefe de Calidad: Panel de Administración", use_container_width=True):
            st.session_state["nav_state"] = "admin_login"
            st.rerun()

# ==========================================
# 2. LOGIN (AMBOS PERFILES)
# ==========================================
elif st.session_state["nav_state"] in ["form_login", "admin_login"]:
    col1, col2, col3 = st.columns([1, 2, 1])
    with col2:
        if st.button("⬅️ Volver al inicio"):
            st.session_state["nav_state"] = "home"
            st.rerun()
        
        es_admin = st.session_state["nav_state"] == "admin_login"
        st.title("Panel de Administrador" if es_admin else "Acceso a Registro")
        st.markdown("Ingrese la contraseña autorizada:")
        
        pwd = st.text_input("Contraseña", type="password")
        pwd_correcta = "glad726lif" if es_admin else "20lf26"
        
        if st.button("Verificar Acceso", use_container_width=True, type="primary"):
            if pwd == pwd_correcta:
                st.session_state["admin_logueado" if es_admin else "form_logueado"] = True
                st.session_state["nav_state"] = "admin_dashboard" if es_admin else "form"
                st.rerun()
            else:
                st.error("Contraseña incorrecta.")

# ==========================================
# 3. FORMULARIO DEL COLABORADOR
# ==========================================
elif st.session_state["nav_state"] == "form":
    if not st.session_state["form_logueado"]:
        st.session_state["nav_state"] = "home"
        st.rerun()
        
    c_h1, c_h2 = st.columns([4, 1])
    with c_h2:
        if st.button("⬅️ Cerrar Sesión", use_container_width=True):
            st.session_state["form_logueado"] = False
            st.session_state["nav_state"] = "home"
            st.rerun()
    
    mostrar_logo(140)
    st.title("Registro de Recepción de Coco")
    
    if st.session_state["enviado_exitoso"]:
        st.success("¡Registro enviado con éxito! Quedará pendiente de validación.")
        if st.button("➕ Ingresar un nuevo registro", type="primary"):
            st.session_state["enviado_exitoso"] = False
            st.rerun()
    else:
        with st.form("form_coco"):
            st.header("1. Datos Generales")
            c1, c2 = st.columns(2)
            with c1:
                responsable = st.selectbox("Nombre del responsable", ["Carlos Canto", "Carlos Rodas", "Jonathan", "Damarias Arellanos", "Carlos López", "Marlon Escobar"])
                proveedor_opcion = st.selectbox("Proveedor", ["GRANOS BASICOS LA PATRONA SOCIEDAD ANONIMA", "Otro"])
                proveedor_final = st.text_input("Si eligió 'Otro', escriba el nombre:") if proveedor_opcion == "Otro" else proveedor_opcion
                desc_materia = st.text_input("Materia prima", value="Coco")
            with c2:
                fecha = st.date_input("Fecha", value=datetime.today())
                hora = st.text_input("Hora (HH:MM)", value=datetime.now().strftime("%H:%M"))
                total_fruta = st.number_input("Total Fruta Ingresada", min_value=0.0, value=0.0)
                cant_unidades = st.number_input("Unidades (Muestra)", min_value=0.0, value=0.0)
            
            observaciones = st.text_area("Observaciones Generales", value="Ninguna")
            
            st.header("2. Parámetros Fisicoquímicos")
            muestras_datos = {}
            for i in range(1, 4):
                st.subheader(f"Muestra {i}")
                mc1, mc2, mc3, mc4 = st.columns(4)
                with mc1: muestras_datos[f"ug_{i}"] = st.number_input(f"Unidades/Galón (M{i})", min_value=0.0, value=0.0)
                with mc2: muestras_datos[f"v_{i}"] = st.number_input(f"Volumen (M{i})", min_value=0.0, value=0.0)
                with mc3: muestras_datos[f"b_{i}"] = st.number_input(f"Brix° (M{i})", min_value=0.0, value=0.0, format="%.2f")
                with mc4: muestras_datos[f"ph_{i}"] = st.number_input(f"pH (M{i})", min_value=0.0, value=0.0, format="%.2f")

            st.header("3. Evidencia Fotográfica")
            st.info("Adjunte una foto de la medición realizada (pH o Brix).")
            evidencia_foto = st.file_uploader("Seleccionar imagen o tomar foto", type=["png", "jpg", "jpeg"])

            submitted = st.form_submit_button("Guardar y Enviar a Revisión", type="primary")
            
            if submitted:
                id_nuevo = generar_id_registro()
                nombre_evidencia = ""
                if evidencia_foto is not None:
                    img = Image.open(evidencia_foto)
                    if img.mode != 'RGB':
                        img = img.convert('RGB')
                    nombre_evidencia = f"evidencia_{id_nuevo}.jpg"
                    img.save(os.path.join(EVIDENCIAS_DIR, nombre_evidencia))

                prov = proveedor_final if proveedor_opcion == "Otro" else proveedor_opcion
                nuevo_registro = {
                    "ID_Registro": id_nuevo,
                    "Estado": "Pendiente",
                    "Responsable": responsable, "Fecha": str(fecha), "Hora": str(hora),
                    "Desc_Materia": desc_materia, "Observaciones": observaciones,
                    "Proveedor": prov, "Total_Fruta": total_fruta, "Cant_Unidades": cant_unidades,
                    "unidades_galon_1": muestras_datos["ug_1"], "volumen_1": muestras_datos["v_1"], "brix_1": muestras_datos["b_1"], "ph_1": muestras_datos["ph_1"],
                    "unidades_galon_2": muestras_datos["ug_2"], "volumen_2": muestras_datos["v_2"], "brix_2": muestras_datos["b_2"], "ph_2": muestras_datos["ph_2"],
                    "unidades_galon_3": muestras_datos["ug_3"], "volumen_3": muestras_datos["v_3"], "brix_3": muestras_datos["b_3"], "ph_3": muestras_datos["ph_3"],
                    "Evidencia": nombre_evidencia,
                    "Firma_Jefe": "Sin firma",
                    "Observaciones_Jefe": ""
                }
                df = cargar_datos()
                df = pd.concat([df, pd.DataFrame([nuevo_registro])], ignore_index=True)
                guardar_datos(df)
                st.session_state["enviado_exitoso"] = True
                st.rerun()

# ==========================================
# 4. DASHBOARD DEL ADMINISTRADOR
# ==========================================
elif st.session_state["nav_state"] == "admin_dashboard":
    if not st.session_state.get("admin_logueado", False):
        st.session_state["nav_state"] = "home"
        st.rerun()
        
    c_head1, c_head2 = st.columns([5, 1])
    with c_head1:
        st.title("Panel de Administrador - Jefe de Calidad")
    with c_head2:
        if st.button("Cerrar sesión"):
            st.session_state["admin_logueado"] = False
            st.session_state["nav_state"] = "home"
            st.rerun()
            
    df = cargar_datos()
    
    total_pen = len(df[df["Estado"] == "Pendiente"]) if not df.empty else 0
    total_apr = len(df[df["Estado"] == "Aprobado"]) if not df.empty else 0
    total_rec = len(df[df["Estado"] == "Rechazado"]) if not df.empty else 0
    total_reg = len(df) if not df.empty else 0
    
    tab_pendientes, tab_aprobados, tab_rechazados, tab_todos = st.tabs([
        f"⏳ Pendientes ({total_pen})", 
        f"✅ Aprobados ({total_apr})", 
        f"❌ Rechazados ({total_rec})", 
        f"📊 Registros ({total_reg})"
    ])
    
    def render_tarjeta(row, index_key, allow_review=False):
        estado_icono = "⏳" if row['Estado'] == "Pendiente" else "✅" if row['Estado'] == "Aprobado" else "❌"
        css_class = f"status-{row['Estado'].lower()}"
        
        st.markdown(f"""
        <div class="record-card">
            <div class="record-header">{estado_icono} #{row['ID_Registro']} — {row['Proveedor']}</div>
            <div class="record-sub">Fecha: {row['Fecha']} | Receptor: {row['Responsable']} | Estado: <span class="{css_class}">{row['Estado']}</span></div>
        </div>
        """, unsafe_allow_html=True)
        
        c_btn1, c_btn2, c_btn3 = st.columns([2, 2, 8])
        
        with c_btn1:
            if st.button("🗑️ Eliminar", key=f"del_{index_key}_{row['ID_Registro']}"):
                eliminar_registro(row['ID_Registro'])
                st.rerun()
                
        with c_btn2:
            if row['Estado'] == 'Aprobado':
                pdf_bytes = generar_pdf_nuevo(row.to_dict())
                st.download_button("📥 PDF", data=pdf_bytes, file_name=f"Recepcion_{row['ID_Registro']}.pdf", mime="application/pdf", key=f"pdf_{index_key}_{row['ID_Registro']}")

        # Espacios invisibles para evitar conflicto de keys en los expanders
        espaciador = " " * (1 if index_key == "pen" else 2 if index_key == "apr" else 3 if index_key == "rec" else 4)
        with st.expander(f"Ver detalles del registro #{row['ID_Registro']}{espaciador}"):
            st.write(f"**Materia Prima:** {row['Desc_Materia']} | **Fruta:** {row['Total_Fruta']} | **Unidades:** {row['Cant_Unidades']}")
            st.write(f"**Observaciones del Colaborador:** {row['Observaciones']}")
            st.write("---")
            
            c_m1, c_m2, c_m3 = st.columns(3)
            with c_m1: st.info(f"**M1:** Galón: {row['unidades_galon_1']} | Vol: {row['volumen_1']} | Brix: {row['brix_1']} | pH: {row['ph_1']}")
            with c_m2: st.info(f"**M2:** Galón: {row['unidades_galon_2']} | Vol: {row['volumen_2']} | Brix: {row['brix_2']} | pH: {row['ph_2']}")
            with c_m3: st.info(f"**M3:** Galón: {row['unidades_galon_3']} | Vol: {row['volumen_3']} | Brix: {row['brix_3']} | pH: {row['ph_3']}")
            
            if "Evidencia" in row and pd.notna(row["Evidencia"]) and row["Evidencia"] != "":
                ruta_evidencia = os.path.join(EVIDENCIAS_DIR, str(row["Evidencia"]))
                if os.path.exists(ruta_evidencia):
                    st.write("**Evidencia Fotográfica:**")
                    st.image(ruta_evidencia, width=350)
            
            if row.get("Observaciones_Jefe", "") != "":
                st.warning(f"**Observaciones de Calidad:** {row['Observaciones_Jefe']}")
            
            if allow_review and row['Estado'] == "Pendiente":
                st.markdown("#### ✍️ Evaluación de Calidad")
                obs_jefe = st.text_area("Añadir observaciones (Opcional):", key=f"obs_jefe_{row['ID_Registro']}")
                
                st.write("**Firma de Aprobación (Dibuje aquí):**")
                
                # Canvas corregido: fondo blanco sólido y update=False para evitar problemas de React
                canvas_result = st_canvas(
                    fill_color="#ffffff",
                    stroke_width=2, 
                    stroke_color="#115e59",
                    background_color="#ffffff",
                    height=120, 
                    width=350, 
                    drawing_mode="freedraw",
                    update_streamlit=False,
                    key=f"canvas_{index_key}_{row['ID_Registro']}"
                )
                
                st.markdown("<br>", unsafe_allow_html=True)
                c_rev1, c_rev2 = st.columns(2)
                with c_rev1:
                    if st.button("✅ Aprobar Registro", key=f"btn_aprobar_{row['ID_Registro']}", type="primary"):
                        if canvas_result.image_data is not None:
                            img = Image.fromarray(canvas_result.image_data.astype('uint8'), mode="RGBA")
                            nombre_firma = f"firma_{row['ID_Registro']}.png"
                            img.save(os.path.join(FIRMAS_DIR, nombre_firma))
                            
                            df.loc[df['ID_Registro'] == row['ID_Registro'], 'Estado'] = "Aprobado"
                            df.loc[df['ID_Registro'] == row['ID_Registro'], 'Firma_Jefe'] = nombre_firma
                            df.loc[df['ID_Registro'] == row['ID_Registro'], 'Observaciones_Jefe'] = obs_jefe
                            guardar_datos(df)
                            st.rerun()
                        else:
                            st.error("Dibuja tu firma en el recuadro para poder aprobar.")
                
                with c_rev2:
                    if st.button("❌ Rechazar Registro", key=f"btn_rechazar_{row['ID_Registro']}"):
                        df.loc[df['ID_Registro'] == row['ID_Registro'], 'Estado'] = "Rechazado"
                        df.loc[df['ID_Registro'] == row['ID_Registro'], 'Observaciones_Jefe'] = obs_jefe
                        guardar_datos(df)
                        st.rerun()

    # --- PESTAÑA: PENDIENTES ---
    with tab_pendientes:
        if not df.empty:
            df_pen = df[df["Estado"] == "Pendiente"]
            for idx, row in df_pen.iterrows():
                render_tarjeta(row, "pen", allow_review=True)
        else:
            st.info("No hay registros pendientes.")

    # --- PESTAÑA: APROBADOS ---
    with tab_aprobados:
        if not df.empty:
            df_apr = df[df["Estado"] == "Aprobado"]
            for idx, row in df_apr.iterrows():
                render_tarjeta(row, "apr")

    # --- PESTAÑA: RECHAZADOS ---
    with tab_rechazados:
        if not df.empty:
            df_rec = df[df["Estado"] == "Rechazado"]
            for idx, row in df_rec.iterrows():
                render_tarjeta(row, "rec")

    # --- PESTAÑA: TODOS Y DESCARGA EXCEL ---
    with tab_todos:
        st.write("### Historial Completo y Descarga")
        if not df.empty:
            excel_bytes = generar_excel_bytes(df)
            st.download_button(
                label=f"📥 Descargar Base de Datos Completa Excel ({len(df)} filas)", data=excel_bytes,
                file_name="Base_Registros_LIF.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", type="primary"
            )
            
            st.markdown("<br>", unsafe_allow_html=True)
            cf1, cf2 = st.columns(2)
            with cf1:
                prov_filtro = st.selectbox("Filtrar por Proveedor:", ["Todos"] + list(df["Proveedor"].unique()), key="filter_prov")
            with cf2:
                fechas_unicas = ["Todas"] + list(df["Fecha"].unique())
                fecha_filtro = st.selectbox("Filtrar por Fecha:", fechas_unicas, key="filter_fecha")
            
            df_mostrar = df
            if prov_filtro != "Todos":
                df_mostrar = df_mostrar[df_mostrar["Proveedor"] == prov_filtro]
            if fecha_filtro != "Todas":
                df_mostrar = df_mostrar[df_mostrar["Fecha"] == fecha_filtro]
            
            if df_mostrar.empty:
                st.warning("No hay registros con esos filtros.")
            else:
                for idx, row in df_mostrar.iterrows():
                    render_tarjeta(row, "tod")
